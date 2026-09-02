import csv
import hashlib
import io
import uuid
from pathlib import Path

from django.conf import settings
from django.http import FileResponse, HttpResponse
from rest_framework.exceptions import NotFound, ValidationError
from rest_framework.response import Response
from django.db import transaction

from apps.core.models import Company
from . import models as m, serializers as s, services as svc
from .workspace import WorkspaceAPI, records
from .security import authorize, features, has_permission
from .catalog import DOCUMENT_FEATURES, DOCUMENT_PERMISSION


REGISTRY = {
    "items": (m.Item,s.ItemSerializer,"item","basic"), "customers": (Company,s.CustomerSerializer,"customer","basic"),
    "suppliers": (m.Supplier,s.SupplierSerializer,"supplier","purchase"), "jobs": (m.Job,s.JobSerializer,"job","work_orders"),
    "documents": (m.Document,s.DocumentSerializer,"invoice","basic"), "employees": (m.Employee,s.EmployeeSerializer,"employee","hrms"),
    "attendance": (m.Attendance,s.AttendanceSerializer,"attendance","attendance_hr"), "expenses": (m.Expense,s.ExpenseSerializer,"expense","expense_management"),
    "payments": (m.Payment,s.PaymentSerializer,"payment","basic"), "payroll": (m.PayrollRun,s.PayrollSerializer,"payroll","payroll"),
    "payroll-results": (m.PayrollResult,s.PayrollResultSerializer,"payroll","payroll"),
    "stock": (m.StockBalance,s.StockSerializer,"stock","inventory"), "stock-movements": (m.StockMovement,s.MovementSerializer,"stock","inventory"),
    "warehouses": (m.Warehouse,s.WarehouseSerializer,"warehouse","inventory"), "tasks": (m.Task,s.TaskSerializer,"task","tasks"),
}


def resource_object(request, resource, pk, action="view"):
    if resource not in REGISTRY: raise NotFound("Resource not found.")
    model, serializer, prefix, feature = REGISTRY[resource]
    obj = model.objects.filter(tenant=request.tenant, pk=pk).first()
    if not obj: raise NotFound("Record not found.")
    if isinstance(obj, m.Document): prefix, feature = DOCUMENT_PERMISSION[obj.kind], DOCUMENT_FEATURES[obj.kind]
    if isinstance(obj, Company): authorize(request, f"{prefix}.{action}", feature=feature)
    else: authorize(request, f"{prefix}.{action}", obj, feature)
    return obj


def safe_cell(value):
    text = str(value if value is not None else "")
    if text.startswith(("=", "+", "@", "\t", "\r")) or (text.startswith("-") and not text[1:].replace(".", "", 1).isdigit()):
        return "'" + text
    return text


class ExportView(WorkspaceAPI):
    def get(self, request):
        resource = request.query_params.get("resource", "")
        if resource not in REGISTRY: raise NotFound("Unknown export resource.")
        model, serializer, prefix, feature = REGISTRY[resource]
        kind = request.query_params.get("kind")
        if model == m.Document:
            if kind not in DOCUMENT_FEATURES: raise ValidationError("Select a document kind.")
            prefix, feature = DOCUMENT_PERMISSION[kind], DOCUMENT_FEATURES[kind]
        authorize(request, prefix + ".export", feature=feature)
        qs = model.objects.filter(tenant=request.tenant).order_by("pk") if model == Company else records(request, model, prefix + ".export", feature)
        if kind: qs = qs.filter(kind=kind)
        data = serializer(qs[:10000],many=True,context={"request":request}).data
        stream = io.StringIO(); writer = csv.writer(stream)
        keys = [f for f in data[0] if not isinstance(data[0][f], (dict,list))] if data else ["id"]
        writer.writerow(keys)
        for row in data: writer.writerow([safe_cell(row.get(k)) for k in keys])
        from apps.core.services import audit
        audit(actor=request.user,tenant=request.tenant,action="erp.data_exported",resource=None,after={"resource":resource,"kind":kind,"rows":len(data)})
        response = HttpResponse("\ufeff"+stream.getvalue(),content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="myraid-{resource}.csv"'
        response["Cache-Control"] = "no-store"
        return response


class AttachmentView(WorkspaceAPI):
    def get(self, request):
        resource = request.query_params.get("resource_type", ""); pk = request.query_params.get("resource_id", "")
        resource_object(request,resource,pk)
        authorize(request,"document.view")
        qs=m.Attachment.objects.filter(tenant=request.tenant,resource_type=resource,resource_id=pk)
        return Response({"results": [{"id":str(x.pk),"name":x.name,"size":x.size,"content_type":x.content_type} for x in qs]})

    @transaction.atomic
    def post(self, request):
        resource=request.data.get("resource_type",""); pk=request.data.get("resource_id","")
        obj=resource_object(request,resource,pk)
        authorize(request,"document.upload")
        if resource=="employees": authorize(request,"employee.view_private",obj,"hrms")
        upload=request.FILES.get("file")
        if not upload or upload.size>10*1024*1024: raise ValidationError("Choose a PDF or image no larger than 10 MB.")
        content=upload.read();suffix=Path(upload.name).suffix.lower()
        allowed=(suffix==".pdf" and content.startswith(b"%PDF-")) or (suffix==".png" and content.startswith(b"\x89PNG\r\n\x1a\n")) or (suffix in (".jpg",".jpeg") and content.startswith(b"\xff\xd8\xff"))
        if not allowed: raise ValidationError("Only genuine PDF, PNG and JPEG files are accepted.")
        if suffix==".pdf" and any(token in content for token in (b"/JavaScript",b"/JS ",b"/Launch",b"/EmbeddedFile")):
            raise ValidationError("Active or embedded PDF content is not accepted.")
        key=f"{request.tenant.pk}/{uuid.uuid4().hex}{suffix}"
        root=Path(getattr(settings,"ERP_PRIVATE_MEDIA_ROOT",settings.BASE_DIR/"private-media"))
        target=root/key;target.parent.mkdir(parents=True,exist_ok=True);target.write_bytes(content)
        attachment=m.Attachment.objects.create(tenant=request.tenant,branch=getattr(obj,"branch",None),created_by=request.user,
            resource_type=resource,resource_id=str(pk),name=Path(upload.name).name[:200],object_key=key,
            content_type={".pdf":"application/pdf",".png":"image/png",".jpg":"image/jpeg",".jpeg":"image/jpeg"}[suffix],size=len(content),checksum=hashlib.sha256(content).hexdigest(),sensitivity="private" if resource in ("employees","payroll","payroll-results") else "business")
        svc.record_event(attachment,request.user,"document.attached")
        return Response({"id":attachment.pk,"name":attachment.name},status=201)


class AttachmentDownloadView(WorkspaceAPI):
    def get(self,request,pk):
        obj=m.Attachment.objects.filter(tenant=request.tenant,pk=pk).first()
        if not obj: raise NotFound()
        resource_object(request,obj.resource_type,obj.resource_id)
        authorize(request,"document.view")
        if obj.resource_type=="employees": authorize(request,"employee.view_private",feature="hrms")
        root=Path(getattr(settings,"ERP_PRIVATE_MEDIA_ROOT",settings.BASE_DIR/"private-media")).resolve()
        target=(root/obj.object_key).resolve()
        if not target.is_relative_to(root) or not target.is_file(): raise NotFound("File unavailable.")
        response=FileResponse(target.open("rb"),as_attachment=True,filename=obj.name,content_type=obj.content_type)
        response["Cache-Control"]="no-store";response["X-Content-Type-Options"]="nosniff"
        svc.record_event(obj,request.user,"document.downloaded")
        return response


def make_pdf(title,company,metadata,headers,rows,totals=None,note=""):
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle
    from xml.sax.saxutils import escape
    output=io.BytesIO();doc=SimpleDocTemplate(output,pagesize=A4,rightMargin=38,leftMargin=38,topMargin=42,bottomMargin=42)
    styles=getSampleStyleSheet();styles.add(ParagraphStyle("SmallERP",fontName="Helvetica",fontSize=9,leading=13))
    p=lambda value:Paragraph(escape(str(value)),styles["SmallERP"])
    story=[Paragraph(escape(company),styles["Title"]),Paragraph(escape(title),styles["Heading1"]),Spacer(1,14)]
    for text in metadata:story.append(p(text))
    story.append(Spacer(1,18))
    grid=Table([[p(x) for x in headers]]+[[p(x) for x in r] for r in rows],repeatRows=1,hAlign="LEFT")
    grid.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#e5efdd")),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#dbe5d3")),("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    story.append(grid)
    for text in totals or []:story.extend([Spacer(1,9),p(text)])
    if note:story.extend([Spacer(1,18),p(note)])
    def footer(c,d):c.setFont("Helvetica",8);c.setFillColor(colors.grey);c.drawString(38,25,"Generated by Myraid ERP · Source version retained");c.drawRightString(A4[0]-38,25,str(d.page))
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
    return output.getvalue()


class PrintView(WorkspaceAPI):
    def get(self,request,resource,pk):
        obj=resource_object(request,resource,pk)
        authorize(request,"document.print")
        entity=obj.kind if isinstance(obj,m.Document) else "payroll" if isinstance(obj,m.PayrollResult) else "job" if isinstance(obj,m.Job) else resource
        template=m.Configuration.objects.filter(tenant=request.tenant,kind="print_template",entity_type=entity,status="published",archived=False).order_by("-version").first()
        template_data=template.definition if template else {}
        template_version=template.version if template else 0
        source_version=int(obj.snapshot.get("posted_version",obj.version)) if isinstance(obj,m.Document) else obj.run.version if isinstance(obj,m.PayrollResult) else obj.version
        cached=m.RenderedDocument.objects.filter(tenant=request.tenant,resource_type=resource,resource_id=str(obj.pk),source_version=source_version,template_version=template_version).select_related("attachment").first()
        root=Path(getattr(settings,"ERP_PRIVATE_MEDIA_ROOT",settings.BASE_DIR/"private-media")).resolve()
        if cached:
            target=(root/cached.attachment.object_key).resolve()
            if target.is_relative_to(root) and target.is_file():
                response=FileResponse(target.open("rb"),as_attachment=True,filename=cached.attachment.name,content_type="application/pdf");response["Cache-Control"]="no-store";return response
        if isinstance(obj,m.Document):
            rows=[[x.description,str(x.quantity.normalize()),x.unit,f"INR {x.rate:,.2f}",f"{x.tax_rate.normalize()}%",f"INR {x.gross:,.2f}"] for x in obj.lines.all()]
            data=make_pdf(template_data.get("heading") or obj.number,request.tenant.name,[obj.kind.replace("_"," ").title(),f"Date: {obj.date} | Due: {obj.due_date or '-'} | State: {obj.status}",f"Party: {obj.customer.name if obj.customer else obj.supplier.name if obj.supplier else '-'}",f"Source version: {source_version} | Reference: {obj.reference or '-'}"],["Description","Qty","Unit","Rate","GST","Final total"],rows,[f"Taxable: INR {obj.taxable:,.2f}",f"GST included: INR {obj.tax:,.2f}",f"Final amount: INR {obj.gross:,.2f}"],template_data.get("terms") or obj.terms or obj.notes)
        elif isinstance(obj,m.PayrollResult):
            if obj.run.status!="finalized":raise ValidationError("Payslips are available after payroll finalization.")
            data=make_pdf(f"Payslip · {obj.run.month:%B %Y}",request.tenant.name,[f"Employee: {obj.employee.name} | ID: {obj.employee.code}",f"Payable days: {obj.payable_days} | Payroll version: {obj.run.version}"],["Component","Type","Amount (INR)"],[[x["name"],x["kind"],x["amount"]] for x in obj.components],[f"Gross earnings: INR {obj.gross:,.2f}",f"Deductions: INR {obj.deductions:,.2f}",f"Net pay: INR {obj.net:,.2f}"])
        elif isinstance(obj,m.Job):
            data=make_pdf(obj.number,request.tenant.name,[obj.name,f"Due: {obj.due_date} | Status: {obj.status}",obj.instructions],["Stage","Planned","Completed","Status"],[[s.name,str(s.planned),str(s.completed),s.status] for s in obj.stages.all()])
        else:raise ValidationError("This record does not have a printable document template.")
        key=f"{request.tenant.pk}/generated/{uuid.uuid4().hex}.pdf";target=(root/key);target.parent.mkdir(parents=True,exist_ok=True);target.write_bytes(data)
        attachment=m.Attachment.objects.create(tenant=request.tenant,branch=getattr(obj,"branch",None),created_by=request.user,resource_type=resource,resource_id=str(obj.pk),name=f"myraid-{entity}-{obj.pk}.pdf",object_key=key,content_type="application/pdf",size=len(data),checksum=hashlib.sha256(data).hexdigest(),sensitivity="private" if isinstance(obj,m.PayrollResult) else "business")
        rendered=m.RenderedDocument.objects.create(tenant=request.tenant,branch=getattr(obj,"branch",None),created_by=request.user,resource_type=resource,resource_id=str(obj.pk),source_version=source_version,template_version=template_version,attachment=attachment)
        svc.record_event(rendered,request.user,"document.generated")
        response=FileResponse(target.open("rb"),as_attachment=True,filename=attachment.name,content_type="application/pdf");response["Cache-Control"]="no-store";return response


IMPORT_FIELDS={"customers":["name","address","gst_no"],"suppliers":["name","contact_name","email","phone","gstin","address"],"items":["sku","name","item_type","unit","hsn_sac","sale_rate","purchase_rate","tax_rate","reorder_level"],"employees":["code","name","designation","joining_date","status"]}


class ImportTemplateView(WorkspaceAPI):
    def get(self,request):
        from openpyxl import Workbook
        resource=request.query_params.get("resource")
        if resource not in IMPORT_FIELDS:raise ValidationError("Choose customers, suppliers, items or employees.")
        _,_,prefix,feature=REGISTRY[resource];authorize(request,prefix+".create",feature=feature)
        wb=Workbook();sheet=wb.active;sheet.title="Import";sheet.append(IMPORT_FIELDS[resource])
        for col in sheet.columns:sheet.column_dimensions[col[0].column_letter].width=24
        stream=io.BytesIO();wb.save(stream)
        response=HttpResponse(stream.getvalue(),content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"]=f'attachment; filename="{resource}-template.xlsx"';return response


class ImportView(WorkspaceAPI):
    def post(self,request):
        from openpyxl import load_workbook
        resource=request.data.get("resource")
        if resource not in IMPORT_FIELDS:raise ValidationError("Unsupported import type.")
        model,serializer,prefix,feature=REGISTRY[resource];authorize(request,prefix+".create",feature=feature)
        upload=request.FILES.get("file")
        if not upload or upload.size>10*1024*1024:raise ValidationError("Upload a CSV or XLSX file up to 10 MB.")
        content=upload.read()
        if upload.name.lower().endswith(".csv"):
            data=list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
        elif upload.name.lower().endswith(".xlsx"):
            import zipfile
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                if sum(x.file_size for x in z.infolist())>50*1024*1024:raise ValidationError("Workbook decompressed size is too large.")
            wb=load_workbook(io.BytesIO(content),read_only=True,data_only=False)
            sheet=wb.active;iterator=sheet.iter_rows(values_only=True);headers=[str(x or "") for x in next(iterator)]
            data=[dict(zip(headers,row)) for row in iterator if any(v is not None for v in row)];wb.close()
        else:raise ValidationError("Use a .csv or .xlsx template.")
        if not 1<=len(data)<=10000:raise ValidationError("Import between 1 and 10,000 rows.")
        rows=[];errors=[]
        for i,row in enumerate(data,2):
            row={k:svc.snapshot(v) for k,v in row.items() if k in IMPORT_FIELDS[resource] and v is not None and v!=""}
            if any(isinstance(v,str) and v.startswith("=") for v in row.values()):errors.append({"row":i,"errors":"Formulas are not allowed."});continue
            check=serializer(data=row,context={"request":request})
            if not check.is_valid():errors.append({"row":i,"errors":check.errors})
            rows.append(row)
        job=m.DataJob.objects.create(tenant=request.tenant,created_by=request.user,kind="import",resource=resource,status="validated" if not errors else "invalid",rows=rows,errors=svc.snapshot(errors),checksum=hashlib.sha256(content).hexdigest())
        return Response({"id":job.pk,"rows":len(rows),"errors":errors,"status":job.status,"preview":rows[:10]})


class ImportCommitView(WorkspaceAPI):
    def post(self,request,pk):
        job=m.DataJob.objects.filter(tenant=request.tenant,pk=pk,created_by=request.user).first()
        if not job:raise NotFound()
        model,serializer,prefix,feature=REGISTRY[job.resource];authorize(request,prefix+".create",feature=feature)
        def execute():
            job.refresh_from_db()
            if job.status=="completed":return {"id":job.pk,"processed":job.processed,"status":job.status}
            if job.status!="validated":raise ValidationError("Fix import validation errors first.")
            for row in job.rows:
                check=serializer(data=row,context={"request":request});check.is_valid(raise_exception=True)
                obj=check.save(tenant=request.tenant,**({"created_by":request.user,"branch":request.branch} if issubclass(model,m.Record) else {}))
                if isinstance(obj,m.Record):svc.record_event(obj,request.user,"data.imported")
            job.processed=len(job.rows);job.status="completed";svc.touch(job);svc.record_event(job,request.user,"data.import_completed")
            return {"id":job.pk,"processed":job.processed,"status":job.status}
        return Response(svc.command(request,f"import:{pk}:commit",execute))
